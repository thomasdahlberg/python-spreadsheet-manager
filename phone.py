import openpyxl
import sys
import os
import re


def main():
    wb = get_excel_doc()
    sheet = get_sheet(wb)
    row_keys = range(2, sheet.max_row)
    header_values = get_header_rows(sheet)
    print("Loading document...")
    document = build_document(sheet, row_keys, header_values)
    header_count = 1
    print('Headers: ')
    for value in header_values:
        print(f"{header_count}){value}  ", end="")
        header_count += 1
    print("")
    while True:
        get_phone_column = input("Select Phone Number Column:")
        if int(get_phone_column) in range(1, header_count):
            print(f"{header_values[int(get_phone_column) - 1]} selected")
            break
    clean_phone(document, header_values[int(get_phone_column) - 1])
    write_sheet(wb, sheet.title, document, header_values)
    print("Process Complete!")
    return 0


def get_excel_doc():
    file_obj = {}
    os.chdir('./')
    files = os.listdir()
    file_map = {}
    count = 1
    for file in files:
        if file[-5:] == '.xlsx':
            file_map[count] = file
            count += 1
    doc_range = range(1, len(file_map)+1)
    print(f"Excel Documents: {file_map}")
    get_doc = input("Pick document number: ")
    while int(get_doc) not in doc_range:
        print("Not a valid document")
        get_doc = input("Pick document number: ")
    file_obj["wb"] = openpyxl.load_workbook(f"{file_map[int(get_doc)]}")
    file_obj["name"] = f"{file_map[int(get_doc)]}"
    return file_obj


def get_sheet(wb):
    sheet_range = range(1, len(wb["wb"].sheetnames)+1)
    print("Sheets: ", end="")
    for i in range(len(wb["wb"].sheetnames)):
        print(f"{i+1}) {wb['wb'].sheetnames[i]}   ", end="")
    print("")
    pick_sheet = input("Pick sheet number: ")
    while int(pick_sheet) not in sheet_range:
        print("Not a valid sheet")
        pick_sheet = input("Pick sheet number: ")
    return wb["wb"][wb["wb"].sheetnames[int(pick_sheet)-1]]


def get_header_rows(sheet):
    headers = []
    for i in range(1, sheet.max_column + 1):
        headers.append(sheet.cell(row=1, column=i).value)
    return headers


def build_document(sheet, keys, header_values):
    doc = []
    for key in keys:
        row_dict = {}
        for i in range(1, sheet.max_column + 1):
            row_dict[header_values[i-1]] = sheet.cell(row=key, column=i).value
        doc.append(row_dict)
    return doc


def clean_phone(document, column):
    for row in document:
        valid_phone_num = re.sub(r"\D", "", row[column])
        row[column] = valid_phone_num


def write_sheet(wb_obj, sheetname, doc_arr, headers):
    if f'{sheetname}-clean-phone' in wb_obj["wb"].sheetnames:
        wb_obj["wb"].remove(wb_obj["wb"][f'{sheetname}-clean-phone'])
    wb_obj["wb"].create_sheet(index=1, title=f"{sheetname}-clean-phone")
    new_sheet = wb_obj["wb"][f"{sheetname}-clean-phone"]
    new_sheet.append(headers)
    for i in range(1, len(doc_arr)):
        new_sheet.append(list(doc_arr[i].values()))
    wb_obj["wb"].save(filename=f"{wb_obj['name']}")


def format_numbers(wb_obj):
    pass


if __name__ == "__main__":
    main()
