import openpyxl
import os
import sys

header_values = []
document = []
deduped_document = []
unique_values = {}
filter_cols = []

def main():
    wb = get_excel_doc()
    sheet = get_sheet(wb)
    row_keys = range(2,sheet.max_row)

    for i in range(1, sheet.max_column + 1):
        header_values.append(sheet.cell(row=1, column=i).value)
    header_count = 1
    valid = ['y', 'Y', 'yes', 'Yes', 'n', 'N', 'no', 'No']
    print('Headers: ')
    for value in header_values:
        print(f"{header_count}){value}  ", end="")
        header_count += 1
    print("")
    
    want_sort = input("Sort rows (y/n)?")
    while want_sort not in valid:
        print("Invalid Response")
        want_sort = input("Sort rows (y/n)?")

    want_dedupe = input("Deduplicate rows (y/n)?")
    while want_dedupe not in valid:
        print("Invalid Response")
        want_dedupe = input("Deduplicate rows (y/n)?")

    criteria = get_criteria(sheet, want_sort, want_dedupe, len(header_values))
    sort_col = None

    if criteria == 1:
        return 2
    if criteria["sort"]:
        sort_col = header_values[criteria["sort"]]
        print('Sorting by: ', sort_col)

    if criteria["dedupe"]:
        for i in range(0, len(criteria["dedupe"])):
            filter_cols.append(header_values[int(criteria["dedupe"][i]) - 1])
        print('Dedupe Criteria: ', filter_cols)


    # Sorting and Deduplicating
    for key in row_keys:
        row_dict = {}
        for i in range(1, sheet.max_column +1):
            row_dict[header_values[i-1]] = sheet.cell(row=key, column=i).value
        document.append(row_dict)
    new_doc = []
    
    if sort_col:
        new_doc = sorted(document, key=lambda x: str(x[sort_col]))
    else:
        new_doc = document
    if len(filter_cols) > 0:
        for d in new_doc:
            criteria = ""
            for i in range(0,len(filter_cols)):
                criteria += str(d[filter_cols[i]])
            if criteria not in unique_values:
                unique_values[criteria] = True
                deduped_document.append(d)
        print('Original length: ', len(new_doc))
        print('Deduped length: ', len(deduped_document))
        write_sheet(wb, sheet.title, deduped_document, 1)
    else:
        write_sheet(wb, sheet.title, new_doc, 0)

    print("Process Complete!")
    return 0

def get_excel_doc():
    file_obj = {}
    os.chdir('./workdocs')
    files = os.listdir()
    file_map = {}
    count = 1
    for file in files:
        if file[-5:] == '.xlsx':
            file_map[count] = file
            count += 1
    doc_range = range(1,len(file_map)+1)
    print(f"Excel Documents: {file_map}")
    get_doc = input("Pick document number: ")
    while int(get_doc) not in doc_range:
        print("Not a valid document")
        get_doc = input("Pick document number: ")
    file_obj["wb"] = openpyxl.load_workbook(f"{file_map[int(get_doc)]}")
    file_obj["name"] = f"{file_map[int(get_doc)]}"
    return file_obj

def get_sheet(wb):
    sheet_range = range(1, len(wb["wb"].sheetnames)+1)
    print("Sheets: ", end="")
    for i in range(len(wb["wb"].sheetnames)):
        print(f"{i+1}) {wb['wb'].sheetnames[i]}   ", end="")
    print("")
    pick_sheet = input("Pick sheet number: ")
    while int(pick_sheet) not in sheet_range:
        print("Not a valid sheet")
        pick_sheet = input("Pick sheet number: ")
    return wb["wb"][wb["wb"].sheetnames[int(pick_sheet)-1]]

def get_criteria(sheet, sort, dedupe, header_len):
    yes = ['y', 'Y', 'yes', 'Yes']
    criteria = { "sort" : None, "dedupe" : None }
    if sort in yes and dedupe in yes:
        get_sort_crit = input("Select sort column number (ex: 1):")
        while int(get_sort_crit) not in range(1, header_len + 1):
            get_sort_crit = input("Select sort column number (ex: 1):")
        get_dedupe_crit = input("Select dedupe criteria by column number(s) (ex: 1,2,3):")
        formatted_dedupe_crit = get_dedupe_crit.split(',')
        criteria["sort"] = int(get_sort_crit) - 1
        criteria["dedupe"] = formatted_dedupe_crit
        return criteria
    elif sort in yes:
        get_sort_crit = input("Select sort column number (ex: 1):")
        while int(get_sort_crit) not in range(1, header_len + 1):
            get_sort_crit = input("Select sort column number (ex: 1):")
        criteria["sort"] = int(get_sort_crit) - 1
        return criteria
    elif dedupe in yes:
        get_dedupe_crit = input("Select dedupe criteria by column number(s) (ex: 1,2,3):")
        formatted_dedupe_crit = get_dedupe_crit.split(',')
        criteria["dedupe"] = formatted_dedupe_crit
        return criteria
    else:
        print("No sort/dedupe criteria selected")
        return 1

def write_sheet(wb_obj, sheetname, doc_arr, deduped):
    if deduped:
        if f'{sheetname}-deduped' in wb_obj["wb"].sheetnames:
            wb_obj["wb"].remove(wb_obj["wb"][f'{sheetname}-deduped'])
        wb_obj["wb"].create_sheet(index=1, title=f"{sheetname}-deduped")
        deduped_sheet = wb_obj["wb"][f"{sheetname}-deduped"]
        deduped_sheet.append(header_values)
        for i in range(1, len(doc_arr)):
            deduped_sheet.append(list(doc_arr[i].values()))
    else:
        if f'{sheetname}-sorted' in wb_obj["wb"].sheetnames:
            wb_obj["wb"].remove(wb_obj["wb"][f'{sheetname}-sorted'])
        wb_obj["wb"].create_sheet(index=1, title=f"{sheetname}-sorted")
        sorted_sheet = wb_obj["wb"][f"{sheetname}-sorted"]
        sorted_sheet.append(header_values)
        for i in range(1, len(doc_arr)):
            sorted_sheet.append(list(doc_arr[i].values()))
    wb_obj["wb"].save(filename=f"{wb_obj['name']}")
    return 0


if __name__ == "__main__":
    main()